from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import datetime, time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.audit_log import AuditLog
from app.models.farm import Farm, UserFarm
from app.models.user import User
from app.schemas.audit import AuditLogFilters, AuditLogPage, AuditLogRecord


def add_audit_log(
    db: Session,
    *,
    user_id: str | None,
    action: str,
    entity: str,
    entity_id: str | None = None,
    farm_id: str | None = None,
    details: dict | None = None,
) -> None:
    payload = json.dumps(details, ensure_ascii=False) if details else None
    row = AuditLog(
        user_id=user_id,
        farm_id=farm_id,
        action=action,
        entity=entity,
        entity_id=entity_id,
        details=payload,
    )
    db.add(row)


# ═══════════════════════════════════════════════════════════════════════════════
# HU015 - Revisión de auditorías del sistema (Sprint 8 - Camilo)
#
# COMO: Administrador del sistema
# QUIERO: consultar la auditoría con filtros por usuario, acción, entidad,
#         finca y rango de fechas
# PARA:   poder rastrear quién modificó qué información y cuándo, y responder
#         ante inconsistencias o reclamos sobre los datos de la finca.
#
# Hasta el Sprint 7 este servicio solo sabía ESCRIBIR auditoría (add_audit_log).
# A partir de aquí también sabe LEERLA de forma filtrada y paginada.
# ═══════════════════════════════════════════════════════════════════════════════

# ¿Qué? Acciones que corresponden al ciclo de sesión del usuario.
# ¿Para qué? Permitir que el administrador las excluya del listado.
# ¿Impacto? Un login genera muchos registros; ocultarlos deja ver únicamente
#           los cambios sobre los datos del negocio.
AUTH_ACTIONS: frozenset[str] = frozenset({
    "login",
    "logout",
    "logout_all_sessions",
    "register",
    "email_verified",
    "password_reset",
})


def get_accessible_farm_ids(db: Session, user_id: str | uuid.UUID) -> list[uuid.UUID]:
    """¿Qué? Devuelve los ids de las fincas activas del usuario.

    COMO: Administrador de una finca
    QUIERO: ver únicamente la auditoría de las fincas donde estoy registrado
    PARA:   que la información de otras fincas permanezca aislada y no se
            filtre información entre organizaciones distintas.

    ¿Impacto? Si el usuario no pertenece a ninguna finca activa, retorna una
              lista vacía y la consulta de auditoría no devolverá registros.
    """
    rows = db.execute(
        select(UserFarm.farm_id).where(
            UserFarm.user_id == user_id,
            UserFarm.is_active.is_(True),
        )
    ).scalars().all()
    return list(rows)


def _build_filter_conditions(
    filters: AuditLogFilters,
    accessible_farm_ids: list[uuid.UUID],
) -> list:
    """¿Qué? Traduce los filtros del schema a condiciones SQLAlchemy.

    COMO: desarrollador que mantiene el módulo
    QUIERO: construir las condiciones en un solo lugar
    PARA:   que el conteo total y la página de resultados usen exactamente los
            mismos criterios y nunca se desincronicen.

    ¿Impacto? El alcance por finca se aplica SIEMPRE, incluso si el usuario
              envía un farm_id al que no tiene acceso.
    """
    conditions: list = []

    # Alcance obligatorio: solo fincas del usuario.
    if filters.farm_id is not None:
        # Si pide una finca puntual, debe estar dentro de las accesibles.
        if filters.farm_id in accessible_farm_ids:
            conditions.append(AuditLog.farm_id == filters.farm_id)
        else:
            # Condición imposible → resultado vacío sin exponer si la finca existe.
            conditions.append(AuditLog.farm_id.is_(None))
            conditions.append(AuditLog.farm_id.is_not(None))
    else:
        conditions.append(AuditLog.farm_id.in_(accessible_farm_ids))

    if filters.user_id is not None:
        conditions.append(AuditLog.user_id == filters.user_id)

    # Coincidencia parcial e insensible a mayúsculas: buscar "creat" encuentra
    # "create_bovine" y "CREATE_FARM".
    if filters.action:
        conditions.append(AuditLog.action.ilike(f"%{filters.action}%"))

    if filters.entity:
        conditions.append(AuditLog.entity.ilike(f"%{filters.entity}%"))

    # Las fechas son inclusivas en ambos extremos: end_date 2026-08-18 incluye
    # todo lo ocurrido ese día hasta las 23:59:59.
    if filters.start_date:
        conditions.append(
            AuditLog.created_at >= datetime.combine(filters.start_date, time.min)
        )

    if filters.end_date:
        conditions.append(
            AuditLog.created_at <= datetime.combine(filters.end_date, time.max)
        )

    if not filters.include_auth_events:
        conditions.append(AuditLog.action.notin_(AUTH_ACTIONS))

    return conditions


def list_audit_logs(
    db: Session,
    *,
    current_user_id: str | uuid.UUID,
    filters: AuditLogFilters,
    limit: int = 100,
    offset: int = 0,
) -> AuditLogPage:
    """¿Qué? Consulta paginada y filtrada de la auditoría.

    COMO: Administrador del sistema
    QUIERO: obtener los registros de auditoría que cumplen mis filtros junto
            con el total disponible
    PARA:   revisar el historial en pantalla de forma ágil y saber cuántos
            registros existen en total antes de exportarlos.

    ¿Impacto? Resuelve el nombre del usuario y de la finca en una sola pasada
              (sin consultas N+1) y ordena del más reciente al más antiguo.
    """
    accessible_farm_ids = get_accessible_farm_ids(db, current_user_id)

    # Sin fincas accesibles no hay nada que auditar: se evita la consulta.
    if not accessible_farm_ids:
        return AuditLogPage(total=0, limit=limit, offset=offset, items=[])

    conditions = _build_filter_conditions(filters, accessible_farm_ids)

    # ¿Qué? Total de registros que cumplen los filtros (sin paginar).
    # ¿Para qué? Que el frontend muestre "X de Y" y calcule las páginas.
    total = db.execute(
        select(func.count()).select_from(AuditLog).where(*conditions)
    ).scalar_one()

    # ¿Qué? LEFT JOIN con usuario y finca.
    # ¿Para qué? Traer el nombre legible del responsable y de la finca.
    # ¿Impacto? Es LEFT (outer) porque el usuario puede haber sido eliminado
    #           (user_id queda en NULL) y el registro de auditoría debe
    #           conservarse igualmente.
    stmt = (
        select(AuditLog, User, Farm)
        .outerjoin(User, User.id == AuditLog.user_id)
        .outerjoin(Farm, Farm.id == AuditLog.farm_id)
        .where(*conditions)
        .order_by(AuditLog.created_at.desc())
        .limit(limit)
        .offset(offset)
    )

    items: list[AuditLogRecord] = []
    for log, user, farm in db.execute(stmt).all():
        items.append(
            AuditLogRecord(
                id=log.id,
                user_id=log.user_id,
                farm_id=log.farm_id,
                action=log.action,
                entity=log.entity,
                entity_id=log.entity_id,
                details=log.details,
                created_at=log.created_at,
                user_email=user.email if user else None,
                user_full_name=f"{user.first_name} {user.last_name}" if user else None,
                farm_name=farm.name if farm else None,
            )
        )

    return AuditLogPage(total=total, limit=limit, offset=offset, items=items)


# ═══════════════════════════════════════════════════════════════════════════════
# HU015 - Tarea 15.2: Exportación filtrada de la auditoría (Sprint 8 - Camilo)
#
# COMO: Administrador del sistema
# QUIERO: descargar en CSV o Excel exactamente los registros de auditoría que
#         estoy viendo con mis filtros aplicados
# PARA:   conservar la evidencia fuera del sistema, adjuntarla a un informe o
#         entregarla a un tercero sin darle acceso a la aplicación.
#
# Nota: la exportación NO reutiliza la paginación de pantalla. Se exporta todo
# lo que cumple los filtros (hasta EXPORT_MAX_ROWS) para que el archivo no
# quede recortado a los 100 registros visibles.
# ═══════════════════════════════════════════════════════════════════════════════

# ¿Qué? Tope de filas por archivo exportado.
# ¿Para qué? Evitar que una exportación sin filtros agote la memoria del servidor.
# ¿Impacto? Si se alcanza el tope, el administrador debe acotar el rango de
#           fechas; el archivo indica cuántos registros se incluyeron.
EXPORT_MAX_ROWS = 10_000

# ¿Qué? Encabezados del archivo, en español y en el orden de lectura natural.
# ¿Para qué? Que el archivo sea entendible por alguien que no conoce la base de datos.
EXPORT_HEADERS = [
    "Fecha y hora",
    "Usuario",
    "Correo",
    "Accion",
    "Entidad",
    "ID entidad",
    "Finca",
    "Detalles",
]


def _record_to_row(record: AuditLogRecord) -> list[str]:
    """¿Qué? Convierte un registro de auditoría en una fila plana de texto.

    COMO: Administrador que abre el archivo descargado
    QUIERO: ver la fecha legible y el nombre del responsable, no identificadores
    PARA:   poder interpretar la evidencia sin consultar la base de datos.

    ¿Impacto? Los valores nulos se muestran como texto vacío o como
              "Usuario eliminado", nunca como "None".
    """
    return [
        record.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        record.user_full_name or "Usuario eliminado",
        record.user_email or "",
        record.action,
        record.entity,
        record.entity_id or "",
        record.farm_name or "",
        record.details or "",
    ]


def collect_logs_for_export(
    db: Session,
    *,
    current_user_id: str | uuid.UUID,
    filters: AuditLogFilters,
) -> list[AuditLogRecord]:
    """¿Qué? Obtiene todos los registros que cumplen los filtros, sin paginar.

    COMO: Administrador exportando evidencia
    QUIERO: que el archivo contenga todos los registros filtrados
    PARA:   que la evidencia esté completa y no limitada a la página que veía
            en pantalla en ese momento.
    """
    page = list_audit_logs(
        db,
        current_user_id=current_user_id,
        filters=filters,
        limit=EXPORT_MAX_ROWS,
        offset=0,
    )
    return page.items


def export_to_csv(records: list[AuditLogRecord]) -> bytes:
    """¿Qué? Genera un archivo CSV con los registros recibidos.

    COMO: Administrador
    QUIERO: un archivo liviano que pueda abrir en cualquier herramienta
    PARA:   filtrar o cruzar los datos con otras fuentes rápidamente.

    ¿Impacto? Se codifica en UTF-8 con BOM (utf-8-sig) porque Excel en Windows
              muestra mal las tildes y las eñes si el archivo no lo lleva.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(EXPORT_HEADERS)
    for record in records:
        writer.writerow(_record_to_row(record))
    return buffer.getvalue().encode("utf-8-sig")


def export_to_excel(records: list[AuditLogRecord]) -> bytes:
    """¿Qué? Genera un archivo Excel (.xlsx) con los registros recibidos.

    COMO: Administrador que presenta la auditoría en una reunión
    QUIERO: un archivo con encabezados resaltados y columnas legibles
    PARA:   poder mostrarlo o imprimirlo sin tener que darle formato a mano.

    ¿Impacto? Usa la misma paleta institucional del módulo de reportes (HU013)
              para que los entregables del sistema se vean consistentes.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Auditoria"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="59930A", end_color="59930A", fill_type="solid")

    sheet.append(EXPORT_HEADERS)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for record in records:
        sheet.append(_record_to_row(record))

    # Anchos pensados para el contenido real de cada columna.
    for column, width in zip(
        ["A", "B", "C", "D", "E", "F", "G", "H"],
        [20, 28, 30, 22, 20, 38, 25, 60],
        strict=False,
    ):
        sheet.column_dimensions[column].width = width

    # Congelar la fila de encabezados para no perderla al desplazarse.
    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
