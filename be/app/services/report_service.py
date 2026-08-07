import io
import uuid
from datetime import date

from fastapi import HTTPException, status
from sqlalchemy import text
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

from app.models.user import User
from app.schemas.report import (
    EconomicReport,
    ExportFormat,
    ProductiveReport,
    ReportCategory,
    ReportFilterParams,
    ReportResponse,
    SanitaryReport,
)


def get_farm_name(db: Session, farm_id: uuid.UUID) -> str:
    result = db.execute(
        text("SELECT name FROM farm WHERE id = :farm_id"),
        {"farm_id": str(farm_id)},
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Farm not found")
    return row


def build_productive_report(
    db: Session, farm_id: uuid.UUID, start_date: date | None, end_date: date | None
) -> ProductiveReport:
    stats = db.execute(
        text("SELECT fn_farm_statistics(:farm_id)"),
        {"farm_id": str(farm_id)},
    ).scalar_one_or_none()

    total_bovines = 0
    males = 0
    females = 0
    avg_weight = None
    if stats:
        s = stats
        if isinstance(s, str):
            import json
            s = json.loads(s)
        if not isinstance(s, dict):
            s = {}
        total_bovines = s.get("total_bovines", 0)
        males = s.get("males", 0)
        females = s.get("females", 0)
        avg_weight = s.get("avg_weight")

    milk_query = "SELECT COALESCE(SUM(quantity_liters), 0) AS total FROM milk_production WHERE farm_id = :farm_id"
    milk_params = {"farm_id": str(farm_id)}
    if start_date:
        milk_query += " AND milking_date >= :start_date"
        milk_params["start_date"] = start_date
    if end_date:
        milk_query += " AND milking_date <= :end_date"
        milk_params["end_date"] = end_date

    total_milk = db.execute(text(milk_query), milk_params).scalar() or 0

    calves = db.execute(
        text("""
            SELECT
                CASE
                    WHEN (CURRENT_DATE - birth_date) <= 90 THEN '0-3 meses'
                    WHEN (CURRENT_DATE - birth_date) <= 180 THEN '3-6 meses'
                    WHEN (CURRENT_DATE - birth_date) <= 365 THEN '6-12 meses'
                    ELSE '12+ meses'
                END AS age_group,
                COUNT(*) AS total
            FROM bovine
            WHERE farm_id = :farm_id AND is_active
            AND birth_date IS NOT NULL
            AND (CURRENT_DATE - birth_date) <= 730
            GROUP BY age_group
        """),
        {"farm_id": str(farm_id)},
    ).all()

    calves_by_age = {}
    for row in calves:
        calves_by_age[row[0]] = row[1]
    total_calves = sum(calves_by_age.values())

    return ProductiveReport(
        total_bovines=total_bovines,
        males=males,
        females=females,
        avg_weight=avg_weight,
        total_milk_liters=float(total_milk),
        avg_milk_per_day=0.0,
        total_calves=total_calves,
        calves_by_age_group=calves_by_age,
    )


def build_sanitary_report(
    db: Session, farm_id: uuid.UUID, start_date: date | None, end_date: date | None
) -> SanitaryReport:
    treatment_query = """
        SELECT
            t.treatment_type,
            COUNT(*) AS total
        FROM treatment t
        WHERE t.farm_id = :farm_id
    """
    treatment_params = {"farm_id": str(farm_id)}
    if start_date:
        treatment_query += " AND t.application_date >= :start_date"
        treatment_params["start_date"] = start_date
    if end_date:
        treatment_query += " AND t.application_date <= :end_date"
        treatment_params["end_date"] = end_date
    treatment_query += " GROUP BY t.treatment_type"

    rows = db.execute(text(treatment_query), treatment_params).all()
    pending = 0
    by_type = {}
    for row in rows:
        by_type[row[0]] = row[1]

    total_treatments = sum(by_type.values())
    active_plans = db.execute(
        text("SELECT COUNT(*) FROM sanitary_plan WHERE farm_id = :farm_id AND is_active"),
        {"farm_id": str(farm_id)},
    ).scalar() or 0

    return SanitaryReport(
        total_treatments=total_treatments,
        active_sanitary_plans=active_plans,
        pending_treatments=pending,
        treatments_by_type=by_type,
    )


def build_economic_report(
    db: Session, farm_id: uuid.UUID, start_date: date | None, end_date: date | None
) -> EconomicReport:
    query = """
        SELECT
            COALESCE(SUM(CASE WHEN er.record_type = 'ingreso' THEN er.amount ELSE 0 END), 0) AS total_income,
            COALESCE(SUM(CASE WHEN er.record_type IN ('egreso', 'gasto') THEN er.amount ELSE 0 END), 0) AS total_expense,
            er.category
        FROM economic_record er
        WHERE er.farm_id = :farm_id
    """
    params = {"farm_id": str(farm_id)}
    if start_date:
        query += " AND er.record_date >= :start_date"
        params["start_date"] = start_date
    if end_date:
        query += " AND er.record_date <= :end_date"
        params["end_date"] = end_date
    query += " GROUP BY er.category, er.record_type"

    rows = db.execute(text(query), params).all()
    total_income = 0.0
    total_expense = 0.0
    income_by_activity = {}
    expense_by_activity = {}

    for row in rows:
        amount = float(row[0] or 0) if row[0] else 0.0
        expense = float(row[1] or 0) if row[1] else 0.0
        total_income += amount
        total_expense += expense
        activity = row[2]
        if amount > 0:
            income_by_activity[activity] = income_by_activity.get(activity, 0) + amount
        if expense > 0:
            expense_by_activity[activity] = expense_by_activity.get(activity, 0) + expense

    return EconomicReport(
        total_income=total_income,
        total_expense=total_expense,
        balance=total_income - total_expense,
        income_by_activity=income_by_activity,
        expense_by_activity=expense_by_activity,
    )


def generate_report(
    db: Session,
    farm_id: uuid.UUID,
    current_user: User,
    params: ReportFilterParams,
) -> ReportResponse:
    farm_name = get_farm_name(db, farm_id)
    cat = params.category

    resp = ReportResponse(
        farm_id=farm_id,
        farm_name=farm_name,
        category=cat,
        start_date=params.start_date,
        end_date=params.end_date,
    )

    if cat is None or cat == ReportCategory.PRODUCTIVE:
        resp.productive = build_productive_report(db, farm_id, params.start_date, params.end_date)
    if cat is None or cat == ReportCategory.SANITARY:
        resp.sanitary = build_sanitary_report(db, farm_id, params.start_date, params.end_date)
    if cat is None or cat == ReportCategory.ECONOMIC:
        resp.economic = build_economic_report(db, farm_id, params.start_date, params.end_date)

    return resp


def export_to_excel(report: ReportResponse) -> bytes:
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="59930A", end_color="59930A", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def write_section(ws, title: str, headers: list[str], rows: list[list]):
        ws.append([title])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=13, color="1B4332")
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        for row_data in rows:
            ws.append(row_data)
            for cell in ws[ws.max_row]:
                cell.border = thin_border
        ws.append([])

    # Productive sheet
    if report.productive:
        ws = wb.active
        ws.title = "Productivo"
        p = report.productive
        write_section(ws, "Resumen Productivo", ["Indicador", "Valor"], [
            ["Total Bovinos", p.total_bovines],
            ["Machos", p.males],
            ["Hembras", p.females],
            ["Peso Promedio", p.avg_weight or 0],
            ["Total Leche (L)", p.total_milk_liters],
            ["Total Terneros", p.total_calves],
        ])
        if p.calves_by_age_group:
            age_rows = [[g, str(c)] for g, c in p.calves_by_age_group.items()]
            write_section(ws, "Terneros por Edad", ["Grupo", "Cantidad"], age_rows)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    # Sanitary sheet
    if report.sanitary:
        ws = wb.create_sheet("Sanitario")
        s = report.sanitary
        write_section(ws, "Resumen Sanitario", ["Indicador", "Valor"], [
            ["Total Tratamientos", s.total_treatments],
            ["Planes Activos", s.active_sanitary_plans],
            ["Pendientes", s.pending_treatments],
        ])
        if s.treatments_by_type:
            type_rows = [[t, str(c)] for t, c in s.treatments_by_type.items()]
            write_section(ws, "Tratamientos por Tipo", ["Tipo", "Cantidad"], type_rows)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    # Economic sheet
    if report.economic:
        ws = wb.create_sheet("Economico")
        e = report.economic
        write_section(ws, "Resumen Economico", ["Indicador", "Valor"], [
            ["Total Ingresos", f"${e.total_income:,.2f}"],
            ["Total Egresos", f"${e.total_expense:,.2f}"],
            ["Balance", f"${e.balance:,.2f}"],
        ])
        if e.income_by_activity:
            inc_rows = [[a, f"${v:,.2f}"] for a, v in e.income_by_activity.items()]
            write_section(ws, "Ingresos por Actividad", ["Actividad", "Monto"], inc_rows)
        if e.expense_by_activity:
            exp_rows = [[a, f"${v:,.2f}"] for a, v in e.expense_by_activity.items()]
            write_section(ws, "Egresos por Actividad", ["Actividad", "Monto"], exp_rows)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_to_pdf(report: ReportResponse) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#1B4332"))
    section_style = ParagraphStyle("Section2", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#59930A"))
    normal = styles["Normal"]

    elements = [Paragraph(f"Reporte - {report.farm_name}", title_style)]
    elements.append(Spacer(1, 12))
    if report.category:
        elements.append(Paragraph(f"Categoria: {report.category.value}", normal))
    if report.start_date:
        elements.append(Paragraph(f"Desde: {report.start_date}", normal))
    if report.end_date:
        elements.append(Paragraph(f"Hasta: {report.end_date}", normal))
    elements.append(Spacer(1, 12))

    GREEN = colors.HexColor("#59930A")
    WHITE = colors.white
    LIGHT_GREEN = colors.HexColor("#E8F5E9")

    def add_table(title: str, headers: list[str], data_rows: list[list]):
        elements.append(Paragraph(title, section_style))
        elements.append(Spacer(1, 6))
        table_data = [headers] + data_rows
        t = Table(table_data, colWidths=[120 * mm, 80 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GREEN),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BACKGROUND", (0, 1), (-1, -1), LIGHT_GREEN),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 12))

    if report.productive:
        p = report.productive
        add_table("Resumen Productivo", ["Indicador", "Valor"], [
            ["Total Bovinos", str(p.total_bovines)],
            ["Machos", str(p.males)],
            ["Hembras", str(p.females)],
            ["Peso Promedio", str(p.avg_weight or 0)],
            ["Total Leche (L)", f"{p.total_milk_liters:.1f}"],
            ["Total Terneros", str(p.total_calves)],
        ])

    if report.sanitary:
        s = report.sanitary
        add_table("Resumen Sanitario", ["Indicador", "Valor"], [
            ["Total Tratamientos", str(s.total_treatments)],
            ["Planes Activos", str(s.active_sanitary_plans)],
            ["Pendientes", str(s.pending_treatments)],
        ])

    if report.economic:
        e = report.economic
        add_table("Resumen Economico", ["Indicador", "Valor"], [
            ["Total Ingresos", f"${e.total_income:,.2f}"],
            ["Total Egresos", f"${e.total_expense:,.2f}"],
            ["Balance", f"${e.balance:,.2f}"],
        ])

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()
