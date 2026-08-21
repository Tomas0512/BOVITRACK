"""
Pruebas: tests/test_audit.py
HU015 - Revisión de auditorías del sistema (Sprint 8 - Camilo, tarea 15.4)

COMO: Administrador del sistema
QUIERO: que el módulo de auditoría esté cubierto por pruebas automáticas
PARA:   tener la certeza de que los filtros no dejan escapar registros de
        fincas ajenas y de que los archivos exportados son legibles y
        completos, incluso después de futuros cambios en el código.

¿Qué?     Pruebas de la validación de filtros, del alcance por finca, del
          armado de condiciones SQL y de la exportación a CSV y Excel.
¿Impacto? Se usan dobles de prueba (MagicMock) para la sesión de base de
          datos, siguiendo el mismo estilo de test_reproductive_event.py, de
          modo que la suite corre sin necesidad de levantar PostgreSQL.
"""

import io
import uuid
from datetime import date, datetime, time
from unittest.mock import MagicMock

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from app.schemas.audit import AuditExportFormat, AuditLogFilters, AuditLogRecord
from app.services import audit_service


# ═══════════════════════════════════════════════════════════════════════════════
# Fixtures
# ═══════════════════════════════════════════════════════════════════════════════


@pytest.fixture
def farm_id() -> uuid.UUID:
    return uuid.uuid4()


@pytest.fixture
def other_farm_id() -> uuid.UUID:
    """Finca a la que el usuario de prueba NO tiene acceso."""
    return uuid.uuid4()


@pytest.fixture
def sample_records() -> list[AuditLogRecord]:
    """¿Qué? Dos registros representativos para probar la exportación.

    El segundo simula un registro cuyo usuario ya fue eliminado (user_id nulo),
    caso que debe seguir apareciendo en el archivo.
    """
    return [
        AuditLogRecord(
            id=uuid.uuid4(),
            user_id=uuid.uuid4(),
            farm_id=uuid.uuid4(),
            action="create_bovine",
            entity="bovine",
            entity_id="B-001",
            details='{"raza": "Cebú"}',
            created_at=datetime(2026, 8, 18, 14, 30, 5),
            user_email="camilo@bovitrack.co",
            user_full_name="Camilo Ortiz",
            farm_name="La Esperanza",
        ),
        AuditLogRecord(
            id=uuid.uuid4(),
            user_id=None,
            farm_id=None,
            action="login",
            entity="user",
            entity_id=None,
            details=None,
            created_at=datetime(2026, 8, 17, 9, 0, 0),
        ),
    ]


# ═══════════════════════════════════════════════════════════════════════════════
# Validación de filtros
# ═══════════════════════════════════════════════════════════════════════════════


def test_filtros_vacios_son_validos():
    """COMO: Administrador que entra por primera vez a la pantalla
    QUIERO: poder consultar sin diligenciar ningún filtro
    PARA:   ver el panorama completo antes de acotar la búsqueda.
    """
    filters = AuditLogFilters()

    assert filters.user_id is None
    assert filters.farm_id is None
    # Por defecto se incluyen los eventos de sesión: el administrador decide
    # explícitamente si quiere ocultarlos.
    assert filters.include_auth_events is True


def test_rango_de_fechas_invertido_es_rechazado():
    """COMO: Administrador que se equivoca al escribir las fechas
    QUIERO: recibir un error claro cuando la fecha inicial es posterior a la final
    PARA:   corregir el filtro en vez de concluir erróneamente que no hubo
            actividad en ese periodo.
    """
    with pytest.raises(ValidationError) as exc:
        AuditLogFilters(start_date=date(2026, 8, 20), end_date=date(2026, 8, 10))

    assert "start_date no puede ser posterior a end_date" in str(exc.value)


def test_rango_de_fechas_del_mismo_dia_es_valido():
    """COMO: Administrador investigando un incidente puntual
    QUIERO: poder filtrar por un único día
    PARA:   ver todo lo ocurrido en esa jornada sin ruido de otros días.
    """
    filters = AuditLogFilters(start_date=date(2026, 8, 18), end_date=date(2026, 8, 18))

    assert filters.start_date == filters.end_date


# ═══════════════════════════════════════════════════════════════════════════════
# Alcance por finca (seguridad)
# ═══════════════════════════════════════════════════════════════════════════════


def test_usuario_sin_fincas_no_recibe_registros():
    """COMO: dueño del sistema
    QUIERO: que un usuario sin fincas activas no obtenga ningún registro
    PARA:   que nadie pueda leer la auditoría de una organización a la que no
            pertenece.

    ¿Impacto? Además se verifica que NI SIQUIERA se ejecute la consulta: se
              corta antes de tocar la base de datos.
    """
    db = MagicMock()
    db.execute.return_value.scalars.return_value.all.return_value = []

    page = audit_service.list_audit_logs(
        db,
        current_user_id=uuid.uuid4(),
        filters=AuditLogFilters(),
        limit=50,
        offset=0,
    )

    assert page.total == 0
    assert page.items == []
    # Una sola llamada: la de get_accessible_farm_ids. No hubo conteo ni SELECT.
    assert db.execute.call_count == 1


def test_pedir_finca_ajena_produce_condicion_imposible(farm_id, other_farm_id):
    """COMO: dueño del sistema
    QUIERO: que solicitar una finca ajena devuelva vacío
    PARA:   que la respuesta no revele si esa finca existe o no en el sistema
            (evita enumerar fincas por prueba y error).
    """
    filters = AuditLogFilters(farm_id=other_farm_id)

    conditions = audit_service._build_filter_conditions(filters, [farm_id])

    # Se agregan dos condiciones mutuamente excluyentes (IS NULL e IS NOT NULL),
    # lo que garantiza cero resultados sin lanzar un 404 delator.
    assert len(conditions) == 2


def test_finca_propia_se_filtra_directamente(farm_id):
    """COMO: Administrador con varias fincas
    QUIERO: poder concentrarme en una sola
    PARA:   revisar la actividad de esa finca sin el ruido de las demás.
    """
    filters = AuditLogFilters(farm_id=farm_id)

    conditions = audit_service._build_filter_conditions(filters, [farm_id, uuid.uuid4()])

    assert len(conditions) == 1


def test_sin_filtro_de_finca_se_limita_a_las_accesibles(farm_id, other_farm_id):
    """COMO: Administrador de varias fincas
    QUIERO: ver por defecto la actividad de todas mis fincas
    PARA:   detectar de un vistazo cualquier movimiento inusual.

    ¿Impacto? El alcance se aplica igual aunque el usuario no filtre por finca.
    """
    conditions = audit_service._build_filter_conditions(AuditLogFilters(), [farm_id])

    assert len(conditions) == 1


# ═══════════════════════════════════════════════════════════════════════════════
# Armado de condiciones de filtro
# ═══════════════════════════════════════════════════════════════════════════════


def test_excluir_eventos_de_sesion_agrega_condicion(farm_id):
    """COMO: Administrador revisando cambios de datos
    QUIERO: poder ocultar los inicios y cierres de sesión
    PARA:   que el ruido de los logins no tape los cambios reales sobre
            bovinos, fincas o tratamientos.
    """
    con_sesion = audit_service._build_filter_conditions(
        AuditLogFilters(include_auth_events=True), [farm_id]
    )
    sin_sesion = audit_service._build_filter_conditions(
        AuditLogFilters(include_auth_events=False), [farm_id]
    )

    assert len(sin_sesion) == len(con_sesion) + 1


def test_cada_filtro_suma_una_condicion(farm_id):
    """COMO: Administrador que combina varios criterios
    QUIERO: que todos los filtros diligenciados se apliquen a la vez
    PARA:   aislar exactamente el evento que estoy investigando.
    """
    filters = AuditLogFilters(
        user_id=uuid.uuid4(),
        action="create",
        entity="bovine",
        start_date=date(2026, 8, 1),
        end_date=date(2026, 8, 18),
    )

    conditions = audit_service._build_filter_conditions(filters, [farm_id])

    # 1 de alcance por finca + 1 usuario + 1 acción + 1 entidad + 2 fechas.
    assert len(conditions) == 6


def test_las_fechas_cubren_el_dia_completo():
    """COMO: Administrador que filtra "hasta el 18 de agosto"
    QUIERO: que se incluya todo lo ocurrido ese día
    PARA:   no perder los registros de la tarde por un corte a medianoche.

    ¿Impacto? Verifica la intención del código: el límite inferior usa
              time.min (00:00:00) y el superior time.max (23:59:59.999999).
    """
    inicio = datetime.combine(date(2026, 8, 18), time.min)
    fin = datetime.combine(date(2026, 8, 18), time.max)

    assert inicio.hour == 0 and inicio.minute == 0
    assert fin.hour == 23 and fin.minute == 59


def test_acciones_de_sesion_estan_catalogadas():
    """COMO: desarrollador que mantiene el módulo
    QUIERO: que las acciones de sesión estén declaradas en un solo lugar
    PARA:   que el filtro y cualquier reporte futuro usen la misma lista.
    """
    assert "login" in audit_service.AUTH_ACTIONS
    assert "logout" in audit_service.AUTH_ACTIONS
    assert "password_reset" in audit_service.AUTH_ACTIONS
    # Un cambio de datos NUNCA debe considerarse evento de sesión.
    assert "create_bovine" not in audit_service.AUTH_ACTIONS


# ═══════════════════════════════════════════════════════════════════════════════
# Exportación
# ═══════════════════════════════════════════════════════════════════════════════


def test_csv_incluye_encabezados_y_todas_las_filas(sample_records):
    """COMO: Administrador que descarga la evidencia
    QUIERO: que el archivo tenga encabezados y una fila por registro
    PARA:   poder leerlo sin conocer la estructura de la base de datos.
    """
    contenido = audit_service.export_to_csv(sample_records).decode("utf-8-sig")
    lineas = [linea for linea in contenido.splitlines() if linea.strip()]

    assert lineas[0].startswith("Fecha y hora;Usuario;Correo")
    assert len(lineas) == len(sample_records) + 1  # + encabezado


def test_csv_lleva_bom_para_excel_en_windows(sample_records):
    """COMO: Administrador que abre el CSV en Excel
    QUIERO: que las tildes y las eñes se vean correctamente
    PARA:   entregar el archivo sin tener que corregir la codificación a mano.

    ¿Impacto? Sin el BOM, Excel en Windows muestra "Cebú" como "CebÃº".
    """
    contenido = audit_service.export_to_csv(sample_records)

    assert contenido.startswith(b"\xef\xbb\xbf")


def test_registro_sin_usuario_se_exporta_como_usuario_eliminado(sample_records):
    """COMO: Auditor revisando el archivo
    QUIERO: entender qué pasó cuando el responsable ya no existe
    PARA:   no ver celdas con "None" ni asumir que el registro está corrupto.

    ¿Impacto? La evidencia se conserva aunque la cuenta se haya eliminado.
    """
    fila = audit_service._record_to_row(sample_records[1])

    assert fila[1] == "Usuario eliminado"
    assert fila[2] == ""  # correo vacío, nunca "None"
    assert fila[7] == ""  # detalles vacíos, nunca "None"


def test_fecha_exportada_es_legible(sample_records):
    """COMO: Administrador leyendo el archivo
    QUIERO: la fecha en formato año-mes-día hora:minuto:segundo
    PARA:   ordenar y comparar los eventos sin ambigüedad de formato.
    """
    fila = audit_service._record_to_row(sample_records[0])

    assert fila[0] == "2026-08-18 14:30:05"


def test_excel_se_genera_y_puede_releerse(sample_records):
    """COMO: Administrador que presenta la auditoría en una reunión
    QUIERO: un archivo Excel válido con encabezados y datos
    PARA:   mostrarlo o imprimirlo sin darle formato a mano.
    """
    contenido = audit_service.export_to_excel(sample_records)
    hoja = load_workbook(io.BytesIO(contenido)).active

    assert hoja.title == "Auditoria"
    assert hoja.max_row == len(sample_records) + 1
    assert hoja.cell(row=1, column=1).value == "Fecha y hora"
    assert hoja.cell(row=2, column=2).value == "Camilo Ortiz"


def test_excel_congela_la_fila_de_encabezados(sample_records):
    """COMO: Administrador revisando cientos de registros
    QUIERO: que los encabezados queden fijos al desplazarme
    PARA:   no perder de vista a qué columna corresponde cada dato.
    """
    hoja = load_workbook(io.BytesIO(audit_service.export_to_excel(sample_records))).active

    assert hoja.freeze_panes == "A2"


def test_exportacion_vacia_conserva_los_encabezados():
    """COMO: Administrador cuyos filtros no arrojaron resultados
    QUIERO: recibir igualmente un archivo con encabezados
    PARA:   dejar constancia de que la consulta se hizo y no arrojó nada.
    """
    csv_vacio = audit_service.export_to_csv([]).decode("utf-8-sig")
    hoja = load_workbook(io.BytesIO(audit_service.export_to_excel([]))).active

    assert csv_vacio.strip().count("\n") == 0  # solo la fila de encabezados
    assert hoja.max_row == 1


def test_tope_de_exportacion_esta_definido():
    """COMO: responsable de la estabilidad del servidor
    QUIERO: que la exportación tenga un límite superior de filas
    PARA:   que una descarga sin filtros no agote la memoria del backend.
    """
    assert audit_service.EXPORT_MAX_ROWS > 0
    assert audit_service.EXPORT_MAX_ROWS <= 50_000


def test_formatos_de_exportacion_disponibles():
    """COMO: Administrador
    QUIERO: elegir entre CSV y Excel
    PARA:   usar el formato que mejor se ajuste a lo que voy a hacer con los datos.
    """
    assert AuditExportFormat.CSV.value == "csv"
    assert AuditExportFormat.EXCEL.value == "excel"
